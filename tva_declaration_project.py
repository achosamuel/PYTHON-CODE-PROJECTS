import openpyxl
import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl.styles import Font, Border, Alignment, PatternFill, Side, numbers
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime
print('\033[2J')
print('\033[H')

# 1- file uploade
# 2- check if the file is an excel file with .xlsx
# start the cleanning:
# 3- there is only one sheet in the file, if it is many ask to delete unimportant sheet and re-upload the file


def file_upload():
    while True:
        window = tk.Tk()
        window.withdraw()
        file_path = filedialog.askopenfilename(
            title='Select a file with .XLSX extension', filetypes=[('Excel files', '*.xlsx')])

        if file_path:
            try:
                uploaded_file = openpyxl.load_workbook(file_path)
                if len(uploaded_file.sheetnames) != 1:
                    print('Select a file with only one sheet, Delete unimportant sheet')
                    continue
                else:
                    print('File uploaded succesfully!!\n')
                    return file_path

            except:
                print('File not supported, import an Excel file')
        else:
            exit()


# file_upload()

# 4- clean column character, capitalize column name and use df.str.contains() to match with the required columns to avoid error from wrong spelling


def columns_formatting():
    uploaded_file = file_upload()

    df = pd.read_excel(uploaded_file)

    df.columns = (
        df.columns
        .str.capitalize()
        .str.strip()
        .str.replace(' ', '_')
    )
# # df.dropna()	Rows with any missing value
# df.dropna(axis=1)	Columns with any missing value
# df.dropna(how='all')	Rows where all values are NaN
    df = df.dropna(how='all')
    df = df.dropna(axis=1, how='all')
    df = df.drop(columns=[col for col in df.columns if (col is None) or (
        str(col).strip() == '') or (str(col).startswith('Unnamed'))])
# 5 - check if all the important columns is presents in the sheet - if not print columns name missing
# IMPORTANT COLUMNS CHECK
    df = df[['Date_facture', 'N°_de_facture', 'Mode_de_payement',
             'Fournisseur', 'Type', 'Credit', 'Debit', 'Taux', 'Tva', 'Date_valeur']]

    df = df.dropna(subset=['Tva'])
    df = df[df['Tva'] != 0]
    df = df.dropna(subset=['N°_de_facture'])

# 5 - check if all the important columns is presents in the sheet - if not print columns name missing
# 6 - select only column required from the sheet and delete the others
    # df['Credit'] = df['Credit'].astype(str).str.replace(',', '.').astype(float)
    # df['Debit'] = df['Debit'].astype(str).str.replace(',', '.').astype(float)
    # df['Taux'] = df['Taux'].astype(str).str.replace(',', '.').astype(float)
    # df['Tva'] = df['Tva'].astype(str).str.replace(',', '.').astype(float)

    # df['Credit'] = pd.to_numeric(df['Credit'].astype(
    #     str).str.replace(',', '.'), errors='coerce')
    # df['Debit'] = pd.to_numeric(df['Debit'].astype(
    #     str).str.replace(',', '.'), errors='coerce')
    # df['Tva'] = pd.to_numeric(df['Tva'].astype(
    #     str).str.replace(',', '.'), errors='coerce')

    df[['Credit', 'Debit', 'Taux', 'Tva']] = df[['Credit', 'Debit',
                                                 'Taux', 'Tva']].apply(pd.to_numeric, errors='coerce')
    df['Date_facture'] = pd.to_datetime(
        df['Date_facture'], errors='coerce', utc=False).dt.date

    df['Date_valeur'] = pd.to_datetime(
        df['Date_facture'], errors='coerce', utc=False).dt.date

    df[['N°_de_facture', 'Mode_de_payement', 'Fournisseur', 'Type']] = df[[
        'N°_de_facture', 'Mode_de_payement', 'Fournisseur', 'Type']].astype(str)

    # Column Mode de paiement formatting
    # VRT RECU//
    # VRT EMIS//
    # ENCAISSEMENT//
    # CHQ 0000370//
    # VIREMENT//
    # CHEQUE//
    # CARTE
    # VIT RECU//
    # VRT RCU//
    # ENVAISSEMENT//
    df.loc[df['Mode_de_payement'].str.contains(
        'vrt|vit|vir', case=False, na=False), 'Mode_de_payement'] = 4
    df.loc[df['Mode_de_payement'].str.contains(
        'caiss|espec|envais|encais', case=False, na=False), 'Mode_de_payement'] = 1
    df.loc[df['Mode_de_payement'].str.contains(
        'cheque|chq', case=False, na=False), 'Mode_de_payement'] = 2
    df.loc[df['Mode_de_payement'].str.contains(
        'prelev|prel', case=False, na=False), 'Mode_de_payement'] = 3
    df.loc[df['Mode_de_payement'].str.contains(
        'effet|eff', case=False, na=False), 'Mode_de_payement'] = 5
    df.loc[df['Mode_de_payement'].str.contains(
        'compensation|compens|compe', case=False, na=False), 'Mode_de_payement'] = 6
    df.loc[df['Mode_de_payement'].str.contains(
        'car', case=False, na=False), 'Mode_de_payement'] = 7
    # check if all mode of paiement got assignment
    for row in range(len(df)):
        row_value = df.iloc[row, 2]
        if not str(row_value).isdigit():
            print(
                f' >> Column name: |mode de paiement| not found in our database, change Value:|{row_value}| in the file\n')

    ###################
    # Create column ICE of companies
    entreprises_ice_if = [
        {"nom": "AHL DESIGN & ENGINEERING MAROC",
            "if": "14408306", "ice": "000031725000032"},
        {"nom": "BATIPLUS", "if": "66108192", "ice": "003581724000027"},
        {"nom": "BP", "if": "1084612", "ice": "001534931000016"},
        {"nom": "CARBO 3S", "if": "37699204", "ice": "002310691000094"},
        {"nom": "CELAYA AMENEGEMENTS", "if": "20787184", "ice": "001928134000088"},
        {"nom": "CFG BANK", "if": "1031055", "ice": "000084581000081"},
        {"nom": "EMIX ENGINEERING", "if": "31847634", "ice": "002169125000091"},
        {"nom": "GREEN WORK CONSULTING", "if": "66144603", "ice": "003611957000088"},
        {"nom": "ILEF", "if": "3300718", "ice": "001524978000057"},
        {"nom": "LUSELEC", "if": "15282601", "ice": "000250720000093"},
        {"nom": "LUSEO MEA", "if": "2222", "ice": "2222"},
        {"nom": "MAGHREB RESTAURATION CONSEIL",
            "if": "3382673", "ice": "001764430000097"},
        {"nom": "MARJANE", "if": "1085086", "ice": "001531173000020"},
        {"nom": "MAROC TELCOM", "if": "3332162", "ice": "001522585000066"},
        {"nom": "MB 2 WORKS", "if": "50259124", "ice": "002785583000052"},
        {"nom": "MLS MECHANICAL", "if": "3379547", "ice": "001534192000084"},
        {"nom": "ORANGE", "if": "1086826", "ice": "001524628000001"},
        {"nom": "PM OUTSOURCING", "if": "3379617", "ice": "001536493000037"},
        {"nom": "REDAL ELECTRICITE", "if": "3315990", "ice": "001562062000023"},
        {"nom": "REDAL EAU", "if": "", "ice": ""},
        {"nom": "RETIS INGENIERIE", "if": "65984292", "ice": "003431681000080"},
        {"nom": "2SS BTP", "ice": "001517503000025", "if": "40126172"},
        {"nom": "AQUABAIN", "ice": "001532876000017", "if": "3346500"},
        {"nom": "AUDIT ET ARBITRAGE INTERNATIONAL",
            "ice": "001572618000094", "if": "3304974"},
        {"nom": "BRICOMA", "ice": "001344200000023", "if": "1087598"},
        {"nom": "COREDIA", "ice": "001909125000007", "if": "20765280"},
        {"nom": "DESKTOP MAROC", "ice": "00230745000001", "if": "1049095"},
        {"nom": "DHL", "ice": "001524108000073", "if": "1085486"},
        {"nom": "ECOWELL", "ice": "001525107000066", "if": "1640138"},
        {"nom": "FREYSSIMA", "ice": "000096373000051", "if": "3382966"},
        {"nom": "IAM", "ice": "001522585000066", "if": "3332162"},
        {"nom": "IMPRIMERIE NAPOLI", "ice": "001549093000016", "if": "3301322"},
        {"nom": "ISOPROJECTS", "ice": "000516718000027", "if": "14490606"},
        {"nom": "KYAGEM", "ice": "001635129000097", "if": "3345702"},
        {"nom": "MATERIA", "ice": "001530465000027", "if": "40147176"},
        {"nom": "OTHENTHIS", "ice": "001698076000065", "if": "3372695"},
        {"nom": "PATHE CALIFORNIE", "ice": "001981172000073", "if": "26041980"},
        {"nom": "SETEC", "ice": "001529590000081", "if": "3330080"},
        {"nom": "SOCIETE D'AMENAGEMENT ZENATA",
            "ice": "001526704000015", "if": "3304707"},
        {"nom": "STE.RABAT INCENDIE", "ice": "000016946000007", "if": "3345612"}
    ]

    df['ICE_FRS'] = ''
    df['IF'] = ''
    print(df.columns.tolist())
    for row in range(len(df)):
        for entreprise in entreprises_ice_if:
            if df.iloc[row, 3].lower().strip() == entreprise['nom'].lower().strip():
                df.iloc[row, 10] = entreprise['ice']
                df.iloc[row, 11] = entreprise['if']

    # df['ICE_FRS'] = df['ICE_FRS'].astype(str)
    ############################
    # Taux column formatting
    df['Taux'] = df['Taux'].apply(lambda x: (str(int(x*100))+"%"))

    # # replace prestation CA par service
    # df.loc[df['Type'].str.contains(
    #     'servic|presta', case=False), 'Type'] = 'SERVICE'
    # df.loc[df['Fournisseur'].str.contains(
    #     'bank|cfg|bce|banque|Attijari|BCP|BMCE |Crédit du Maroc|Société Générale|BMCI', case=False), 'Type'] = 'SERVICE BANCAIRE'

    ######################
    # Tva remove '-'
    df['Tva'] = df['Tva'].abs()

    # credit excel file creating
    credit_df = df[df['Credit'].notna() & (df['Credit'] != '')]
    credit_df['M_HT'] = (credit_df['Credit'] - credit_df['Tva'])
    credit_df.drop(['Debit'], axis=1, inplace=True)
    credit_df.rename(columns={'Credit': 'M_TTC'},
                     inplace=True)  # rename column credit

    credit_df.to_excel('Encaissement_df.xlsx', engine='openpyxl',
                       index=False)  # credit file save

    # Debit excel file creating
    Debit_df = df[df['Debit'].notna() & (df['Debit'] != '')]
    Debit_df.drop(['Credit'], axis=1, inplace=True)
    Debit_df['M_HT'] = (Debit_df['Debit'] - Debit_df['Tva'])
    Debit_df.rename(columns={'Debit': 'M_TTC'},
                    inplace=True)   # rename column Debit

    # replace prestation CA par service
    Debit_df.loc[Debit_df['Type'].str.contains(
        'servic|presta', case=False), 'Type'] = 'SERVICE'
    Debit_df.loc[Debit_df['Fournisseur'].str.contains(
        'bank|cfg|bce|banque|Attijari|BCP|BMCE |Crédit du Maroc|Société Générale|BMCI', case=False), 'Type'] = 'SERVICE BANCAIRE'

    Debit_df.to_excel('Deduction.xlsx', engine='openpyxl',
                      index=False)  # Debit file save

    # df.to_excel('transformed_file.xlsx', index=False)

    # CREATE the file that match the period we are interesting about
    # credit_df['Period'] = credit_df['Date_facture'].dt.strftime('%Y-%m')
    credit_df['Period'] = credit_df['Date_facture'].apply(lambda x: str(x)[:7])
    activate_function = input('Do you want to calculate credit tax(y/n):')
    if activate_function.isalpha and activate_function.lower()[0] == 'y':
        while True:
            year = input(
                'Enter the Year you are interrested by(exemple: 2019): ')
            if year.isdigit and len(str(year)) == 4:
                print('valid character!')
                break
            else:
                print('Invalid year, type again\n')

        while True:
            month = input('Enter the month you are interrested by(1 - 12): ')
            if month.isdigit and 1 <= int(month) <= 12:
                print('Valid input!')
                break
            else:
                print('Invalid input, type again\n')

        period = str(year)+'-'+str(month)
        Credit_filename = f'Encaissement_{year}_{month}'
        file = credit_df[credit_df['Period'] == period]
        file.drop(['Period'], axis=1, inplace=True)
        file.sort_values(['Date_facture'], ascending=True, inplace=True)
        file.to_excel(f'{Credit_filename}.xlsx',
                      engine='openpyxl', index=False)
    else:
        print(f'\n ok lets move to the next')

    # CREATE the file that match the period we are interesting about
    # Debit_df['Period'] = Debit_df['Date_facture'].dt.strftime('%Y-%m')
    Debit_df['Period'] = Debit_df['Date_facture'].apply(lambda x: str(x)[:7])
    activate_function = input('Do you want to calculate debit tax(y/n):')
    if activate_function.isalpha and activate_function.lower()[0] == 'y':
        while True:
            year = input(
                'Enter the Year you are interrested by(exemple: 2019): ')
            if year.isdigit and len(str(year)) == 4:
                print('valid character!')
                break
            else:
                print('Invalid year, type again\n')

        while True:
            month = input('Enter the month you are interrested by(1 - 12): ')
            if month.isdigit and 1 <= int(month) <= 12:
                print('Valid input!')
                break
            else:
                print('Invalid input, type again\n')

        period = str(year)+'-'+str(month)
        Debit_filename = f'Deduction_{year}_{month}'
        file = Debit_df[Debit_df['Period'] == period]
        file.drop(['Period'], axis=1, inplace=True)
        file.sort_values(['Date_facture'], ascending=True, inplace=True)
        file.to_excel(f'{Debit_filename}.xlsx', engine='openpyxl', index=False)

    return Credit_filename, Debit_filename


def separate_file():
    df = pd.read_excel('transformed_file.xlsx')
    # df['ICE_FRS'] = df['ICE_FRS'].fillna('')
    # df['IF'] = df['IF'].fillna('')
    # df['ICE_FRS'] = df['ICE_FRS'].astype(str)
    # #########
    # df['ICE_FRS'] = df['ICE_FRS'].astype(str).str.zfill(15)  # or .str.strip()
    # df['IF'] = df['IF'].astype(str).str.strip()
    df['ICE_FRS'] = df['ICE_FRS'].apply(
        lambda x: str(x).zfill(15) if pd.notnull(x) else '')
    df['IF'] = df['IF'].apply(lambda x: str(
        x).strip() if pd.notnull(x) else '')
    # df['ICE_FRS'] = df['ICE_FRS'].apply(
    #     lambda x: '{:.0f}'.format(x) if pd.notnull(x) else '')

    # df['IF'] = df['IF'].apply(
    #     lambda x: '{:.0f}'.format(x) if pd.notnull(x) else '')

    df[['Credit', 'Debit', 'Tva']] = df[['Credit', 'Debit', 'Tva']].apply(
        pd.to_numeric, errors='coerce')
    df['Date_facture'] = pd.to_datetime(
        df['Date_facture'], errors='coerce', utc=False).dt.date
    df['Date_valeur'] = pd.to_datetime(
        df['Date_valeur'], errors='coerce', utc=False).dt.date

    df[['N°_de_facture', 'Mode_de_payement', 'Fournisseur', 'Type', 'Taux', 'ICE_FRS', 'IF']] = df[[
        'N°_de_facture', 'Mode_de_payement', 'Fournisseur', 'Type', 'Taux', 'ICE_FRS', 'IF']].astype(str)
    ##########

    credit_df = df[df['Credit'].notna() & (df['Credit'] != '')]
    credit_df['M_HT'] = (credit_df['Credit'] - credit_df['Tva'])
    credit_df.drop(['Debit'], axis=1, inplace=True)
    # credit_df['ICE_FRS'] = credit_df['ICE_FRS'].apply(
    #     lambda x: '{:.0f}'.format(x) if pd.notnull(x) else '')
    # credit_df['ICE_FRS'] = credit_df['ICE_FRS'].fillna('')
    # credit_df['IF'] = credit_df['IF'].fillna('')
    # credit_df['ICE_FRS'] = credit_df['ICE_FRS'].astype(str).str.zfill(15)

    Debit_df = df[df['Debit'].notna() & (df['Debit'] != '')]
    Debit_df.drop(['Credit'], axis=1, inplace=True)
    Debit_df['M_HT'] = (Debit_df['Debit'] - Debit_df['Tva'])
    Debit_df['ICE_FRS'] = Debit_df['ICE_FRS'].fillna('')
    Debit_df['IF'] = Debit_df['IF'].fillna('')
    Debit_df['ICE_FRS'] = Debit_df['ICE_FRS'].astype(str).str.zfill(15)

    credit_df.to_excel('credit_df.xlsx', engine='openpyxl', index=False)
    Debit_df.to_excel('Debit_df.xlsx', engine='openpyxl', index=False)

# 7 - separate deduction and encaissement TVA in two different excel files
# 8 - separate by month each file by month


def calculate_credit_declaration():
    credit_df = pd.read_excel('Encaissement_df.xlsx')

    credit_df['Period'] = credit_df['Date_facture'].dt.strftime('%Y-%m')
    activate_function = input('Do you want to calculate credit tax(y/n):')
    if activate_function.isalpha and activate_function.lower()[0] == 'y':
        while True:
            year = input(
                'Enter the Year you are interrested by(exemple: 2019): ')
            if year.isdigit and len(str(year)) == 4:
                print('valid character!')
                break
            else:
                print('Invalid year, type again\n')

        while True:
            month = input('Enter the month you are interrested by(1 - 12): ')
            if month.isdigit and 1 <= int(month) <= 12:
                print('Valid input!')
                break
            else:
                print('Invalid input, type again\n')

        period = str(year)+'-'+str(month)
        filename = f'Encaissement_{year}_{month}'
        file = credit_df[credit_df['Period'] == period]
        file.to_excel(f'{filename}.xlsx', engine='openpyxl', index=False)
        return filename
    else:
        return


def calculate_debit_declaration():
    debit_df = pd.read_excel('Debit_df.xlsx')
    debit_df['Period'] = debit_df['Date_facture'].dt.strftime('%Y-%m')
    activate_function = input('Do you want to calculate debit tax(y/n):')
    if activate_function.isalpha and activate_function.lower()[0] == 'y':
        while True:
            year = input(
                'Enter the Year you are interrested by(exemple: 2019): ')
            if year.isdigit and len(str(year)) == 4:
                print('valid character!')
                break
            else:
                print('Invalid year, type again\n')

        while True:
            month = input('Enter the month you are interrested by(1 - 12): ')
            if month.isdigit and 1 <= int(month) <= 12:
                print('Valid input!')
                break
            else:
                print('Invalid input, type again\n')

        period = str(year)+'-'+str(month)
        filename = f'Debit_{year}_{month}'
        file = debit_df[debit_df['Period'] == period]
        file.to_excel(f'{filename}.xlsx', engine='openpyxl', index=False)
    else:
        return

# wb = openpyxl.load_workbook('transformed_file.xlsx')
# ws = wb.active

# for row in range(1, 10):
#     for col in range(1, 15):
#         data = ws.cell(row=row, column=col).value
#         if data is None:
#             data = ''
#         print(f'{str(data):<35}', end=' | ')
#     print()


# 9 - create a function that make deduction declaration automatically
# 10 - create a function that make encaissement declaration automatically
# 11 - create a new file just for the format conversion of the files without delete the old version
# 12 - show button to the user and ask him to download the file i want or interrested by

def Deduction_declaration():
    Credit_filename, Debit_filename = columns_formatting()
    old_excel = openpyxl.load_workbook(f'{Debit_filename}.xlsx')
    new_excel = openpyxl.Workbook()

    # RAISON SOCIAL		MAROC LUSEO
# ID_FISCAL		4101887
# ANNEE		2025
# PERIODE(Mois)		6
# REGIME(Encais-1)		1

    thin = Side(border_style='thin', color='000000')

    ws_new_excel = new_excel.active
    ws_new_excel.title = 'EDI'
    ws_new_excel.sheet_view.showGridLines = False
    ws_new_excel.sheet_properties.tabColor = "0070C0"
    ws_new_excel['A2'] = 'RAISON SOCIAL'
    ws_new_excel['A3'] = 'ID_FISCAL'
    ws_new_excel['A4'] = 'ANNEE'
    ws_new_excel['A5'] = 'PERIODE(Mois)'
    ws_new_excel['A6'] = 'REGIME(Encais-1)'

    for row in range(2, 7):
        start_row = f'A{row}'
        end_row = f'B{row}'
        # fonts
        ws_new_excel[start_row].font = Font(bold=True, size=9)
        # boders
        ws_new_excel[start_row].border = Border(
            left=thin, top=thin, bottom=thin, right=thin)
        # merges cells
        ws_new_excel.merge_cells(f'{start_row}:{end_row}')

    ws_new_excel = new_excel.active
    ws_new_excel['C2'] = input('Enter the company name: ').upper()
    ws_new_excel['C3'] = input('Enter the company ID_FISCAL: ')
    ws_new_excel['C4'] = input('Enter the year: ')
    ws_new_excel['C5'] = input('Enter the PERIODE(Mois): ')
    ws_new_excel['C6'] = input('Enter the REGIME(Encais-1): ')

    for row in range(2, 7):
        cell = f'C{row}'
        ws_new_excel[cell].border = Border(
            left=thin, top=thin, bottom=thin, right=thin)
        ws_new_excel[cell].alignment = Alignment(horizontal='right')
        ws_new_excel[cell].font = Font(size=9)

    ws_new_excel['E5'] = 'Relevé de déduction'
    ws_new_excel['E5'].font = Font(bold=True, size=11)
    ws_new_excel['E5'].border = Border(
        left=thin, top=thin, bottom=thin, right=thin)
    ws_new_excel.merge_cells('E5:J5')
    ws_new_excel['E5'].alignment = Alignment(horizontal='center')

    ws_new_excel['E6'] = '(Article 112 du Code Général des Impôts)'
    ws_new_excel['E6'].border = Border(
        left=thin, bottom=thin, right=thin)
    ws_new_excel.merge_cells('E6:J6')
    ws_new_excel['E6'].alignment = Alignment(horizontal='center')

    ws_new_excel['L1'] = 'Modèle n° ADC082F-15I'
    ws_new_excel['L1'].font = Font(size=8)
    ws_new_excel['L1'].alignment = Alignment(horizontal='center')
    ws_new_excel.merge_cells('L1:M1')

    amoirie_image = Image('TVA DECLARATION FOLDER/amoirie.png')
    dgi_logo = Image(
        'TVA DECLARATION FOLDER\direction-generale-des-impots-logo.png')
    amoirie_image.width = 100
    amoirie_image.height = 50

    dgi_logo.width = 100
    dgi_logo.height = 90

    ws_new_excel.add_image(amoirie_image, 'H2')
    ws_new_excel.add_image(dgi_logo, 'M2')

    #################################################################################################################
    ############# tABLE #####################
    ws_old_excel = old_excel.active
    all_row_credit = []
    for row in range(2, ws_old_excel.max_row + 1):
        row_credit = {}
        row_credit['OR'] = row - 1
        row_credit['FACT_NUM'] = ws_old_excel.cell(row=row, column=2).value
        row_credit['DESIGNATION'] = ws_old_excel.cell(row=row, column=5).value
        row_credit['M_HT'] = ws_old_excel.cell(row=row, column=12).value
        row_credit['TVA'] = ws_old_excel.cell(row=row, column=8).value
        row_credit['M_TTC'] = ws_old_excel.cell(row=row, column=6).value
        row_credit['IF'] = ws_old_excel.cell(row=row, column=11).value
        row_credit['LIB_FRSS'] = ws_old_excel.cell(row=row, column=4).value
        row_credit['ICE_FRS'] = ws_old_excel.cell(row=row, column=10).value
        row_credit['TAUX'] = ws_old_excel.cell(row=row, column=7).value
        row_credit['ID_PAIE'] = ws_old_excel.cell(row=row, column=3).value
        row_credit['DATE_PAIE'] = ws_old_excel.cell(row=row, column=9).value
        row_credit['DATE_FAC'] = ws_old_excel.cell(row=row, column=1).value

        all_row_credit.append(row_credit)

    headers = list(all_row_credit[0].keys())
    for col in range(len(headers)):
        ws_new_excel.cell(row=8, column=col+1).value = headers[col]

    for row, data in enumerate(all_row_credit, start=9):
        for i, col in enumerate(data, start=1):
            ws_new_excel.cell(row=row, column=i).value = data[col]

    # AJUST the cell width
    for column in ws_new_excel.iter_cols():
        max_length = max(len(str(cell.value))
                         for cell in column[7:] if cell.value)
        ws_new_excel.column_dimensions[column[9]
                                       .column_letter].width = max_length + 5

    # FORMATTING
    for row in range(9, ws_new_excel.max_row+1):
        for col in range(1, ws_new_excel.max_column+1):
            column = get_column_letter(col)
            cell = f'{column}{row}'
            ws_new_excel[cell].font = Font(size=10)
            ws_new_excel[cell].alignment = Alignment(horizontal='center')

    # TABLE FORMATTING
    cells = f'A8:M{ws_new_excel.max_row}'
    table = Table(displayName='TABLEAU1', ref=cells)
    style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,
                           showColumnStripes=False, showRowStripes=True, showLastColumn=False)

    table.tableStyleInfo = style
    ws_new_excel.add_table(table)

    for row in ws_new_excel.iter_rows(min_row=8, max_row=ws_new_excel.max_row, min_col=12, max_col=13):
        for cell in row:
            if isinstance(cell.value, datetime):
                cell.number_format = 'DD-MMM-YY'
    # FREEZE Columns heading
    ws_new_excel.freeze_panes = 'A9'

    new_excel.save(f'declaration_{Debit_filename}.xlsx')


# columns_formatting()
# separate_file()
# calculate_credit_declaration()
# calculate_debit_declaration()
Deduction_declaration()
